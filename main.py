import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx") #open the excel file
product_list = inv_file['Sheet1']


product_per_supplier = {}
product_und_10_inv = {}
total_value_per_supplier = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)



    #number if products per supplier
    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier[supplier_name]
        product_per_supplier[supplier_name ] = current_num_products + 1
    else:
        print("adding a new supplier")
        product_per_supplier[supplier_name] = 1

        # total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total = total_value_per_supplier[supplier_name]
        total_value_per_supplier[supplier_name] =  current_total + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    if inventory < 10:
        product_und_10_inv[int(product_num)] = int(inventory)

    #add changes to file
    inventory_price.value = inventory * price


    #save file


print(product_per_supplier)
print(total_value_per_supplier)
print(product_und_10_inv)

inv_file.save("inventory_worked_on.xlsx")










